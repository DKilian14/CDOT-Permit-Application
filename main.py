from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time
import datetime
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from selenium.webdriver.common.by import By
import os
import chromedriver_autoinstaller

date = datetime.datetime.now()

Name=input('enter your name:  ')
Email=input('enter your email:  ')
Phone=input('enter your phone number:  ')
UserName = input("enter your \'City of Chicago\' Permitting USERNAME:  ")
Password = input("enter your \'City of Chicago\' Permitting PASSWORD:  ")

Excel_File = r"excel_files/read/Applications_To_Apply_For.xlsx"

wb = openpyxl.load_workbook(Excel_File)
Excel_Sheet = wb['Sheet1']

PN_Column = "A"
CDOTID_Column = "B"
checklist ='C'
Checklist_Header = (checklist + '1')
Excel_Sheet[''.join(Checklist_Header)] = "CDOT Permit Application Number"
# first_project =
# wb.close()
wb.save("excel_files/read/Applications_To_Apply_For.xlsx")

def find_first_project():
    #If the row in the c column is empty, assign the first project.
    a_num_column = Excel_Sheet[checklist]
    for k in a_num_column:
        if k.value is None:
            break
    return k.row-1



# finds the number of values in the Project Name column
d = Excel_Sheet[PN_Column]
for i in d:
    print(i.value)
    if i.value is None:
        break

# finds the number of values in the CDOT ID column
d = Excel_Sheet[CDOTID_Column]
for c in d:
    print(c.value)
    if c.value is None:
        break
Total_Projects = c.row - 1
print("total projects = ",Total_Projects)

first_project = find_first_project()


# Checks to make sure there are the same number of 'Project Name' values and 'CDOT ID' values.
if c.row - 1 != i.row - 1:
    print('There are either more CDOT IDs or more \'Project Name\'s. '
          'Double check to make sure there are no blank cells in those columns and they are lined up correctly. '
          'Then, restart this program.')
    exit()

chromedriver_autoinstaller.install()

# if there are the same amount of Project Names and CDOT IDs in the excel file,
# a new column will be created to input the DOT Permit #


for i in range(first_project,Total_Projects+1):

    # Finds the coordinates of the next 'Project Name'
    Projectcell_string= PN_Column + str(i + 1)
    Current_Project_Cell = coordinate_from_string(Projectcell_string)
    col = column_index_from_string(Current_Project_Cell[0])
    row = Current_Project_Cell[1]

    # Finds the coordinates of the next 'CDOT ID'
    CDOTITcell_string = CDOTID_Column + str(i + 1)
    Current_CDOTID_Cell = coordinate_from_string(CDOTITcell_string)
    colc = column_index_from_string(Current_CDOTID_Cell[0])
    rowc = Current_CDOTID_Cell[1]

    print("PROJECT NAME: ",Excel_Sheet.cell(row,col).value, ", CDOT ID: ", Excel_Sheet.cell(rowc,colc).value)

    #begin Navigation of CDOT Website

    # Stop website from opening while working on Excel opening and manipulation
    time.sleep(2)
    #C:\Users\danki\PycharmProjects\ChicagoPermitApplications\venv\Lib\site-packages\Chromedriver 98
    s = Service('C:/Users/danki/projects/Chicago_Permit_Application/venv/Lib/site-packages/Chromedriver_autoinstaller/100\chromedriver.exe')
    driver = webdriver.Chrome(service=s)
    driver.get("https://ipi.cityofchicago.org/")
    # Chromedriver Path : C:/Users/danki/PycharmProjects/ChicagoPermitApplications/venv/Lib/site-packages/chromedriver_py\chromedriver.exe

    # Inserting Username
    id_box = driver.find_element(By.NAME,'UserName')
    id_box.send_keys(UserName)

    # Inserting Password
    pass_box = driver.find_element(By.NAME,'Password')
    pass_box.send_keys(Password)
    time.sleep(2)

    # Find login button
    login_button = driver.find_element(By.CSS_SELECTOR,'.btn.btn-primary.btn-full')
    # Click login
    login_button.click()
    time.sleep(2)

    # find "Create New v" button
    create_new = driver.find_element(By.CSS_SELECTOR,'.btn.btn-success.btn-full.dropdown-toggle')

    # click the button
    create_new.click()
    time.sleep(.5)

    # find the "permit application" button
    permit_application = driver.find_element(By.LINK_TEXT,"Permit Application")
    permit_application.click()

    # find the "CDOT DAS or Conduit Periodic Application Process" button
    cdot_das = driver.find_element(By.ID,"formdot_period_dasconduit")
    cdot_das.click()

    # find 'Project Name' field and actualize it
    project_name = driver.find_element(By.NAME,'ApplicationName')
    project_name.send_keys(Excel_Sheet.cell(row,col).value)

    #find button
    type_of_periodic = driver.find_element(By.NAME,'WorkType')
    type_of_periodic.click()

    # from dropdown menu of "type of Periodic", select "DAS-Distributed Antenna System".
    DAS = driver.find_element(By.XPATH,'//*[@id="WorkType"]/option[3]')
    DAS.click()

    # find "Description of Work" field
    description_of_work = driver.find_element(By.NAME,'Comments')
    description_of_work.send_keys('DAS Colocation on CDOT pole.')

    # find "Next Step" button
    next_step = driver.find_element(By.ID,'btnSave')
    next_step.click()

    # find and click "+Add Emergency Contact Information
    em_contact = driver.find_element(By.CSS_SELECTOR,'.btn.btn-success.grid-add-button')
    em_contact.click()

    # Insert "Name" of emergency contact
    em_name = driver.find_element(By.NAME,
        'EmergCntcTdetailpagedelimiterChicagoDOTUseCoreEmergCntcGcontroldelimiterName')
    em_name.send_keys(Name)

    # insert "Phone" of emergency contact
    em_name = driver.find_element(By.NAME,
        'EmergCntcTdetailpagedelimiterChicagoDOTUseCoreEmergCntcGcontroldelimiterPhone')
    em_name.send_keys(Phone)

    # insert 'Email' of emergency contact
    em_email = driver.find_element(By.NAME,
        'EmergCntcTdetailpagedelimiterChicagoDOTUseCoreEmergCntcGcontroldelimiterEmail')
    em_email.send_keys(Email)

    # click the "Add Emergency Contact Information" button
    add_em_button = driver.find_element(By.CSS_SELECTOR,'.btn.btn-primary')
    add_em_button.click()

    # Click the "Next Step" Button
    next_step = driver.find_element(By.ID,'finalSubmit')
    next_step.click()

    # obtain name of CDOT Application
    CDOT_ID_Location = driver.find_element(By.XPATH,'/html/body/div[1]/section/section[2]/div[1]/h4').text
    CDOT_ID = CDOT_ID_Location[20:]
    print("CDOT ID : ",CDOT_ID)

    #inserting the Chicago-generated CDOT ID into the Excel Sheet
    CDOT_ID_Cell = ''.join(checklist + str(i+1))
    Excel_Sheet[CDOT_ID_Cell] = CDOT_ID
    wb.save("excel_files/read/Applications_To_Apply_For.xlsx")

    # Insert "CBI ID" on "Pole Selection" page
    CBI_ID = driver.find_element(By.NAME,'DASInfoTdetailpagedelimitercontroldelimiterEBDASInfoT_PoleFileNumber')
    CBI_ID.send_keys(Excel_Sheet.cell(rowc, colc).value)

    # Click dropdown for "Application Year"
    app_year = driver.find_element(By.NAME,'DASInfoTdetailpagedelimitercontroldelimiterEBDASInfoT_ApplicationYear')
    app_year.click()

    # Click on "current year" from "Application Year" dropdown menu
    current_year = driver.find_element(By.XPATH,
        '//*[@id="DASInfoTdetailpagedelimitercontroldelimiterEBDASInfoT_ApplicationYear"]/option[2]')
    current_year.click()

    # Click "Next Step"
    next_step = driver.find_element(By.ID,'finalSubmit')
    next_step.click()

    # Check the checkbox "I Agree"
    i_agree = driver.find_element(By.NAME,
        'SubmissionRecordTdetailpagedelimiterEditChicagoCoCUseSubmissionRecordGcontroldelimiterWebSubmittedgriditemdelimiter0')
    i_agree.click()

    """

    # FINAL SUBMIT!!! DO NOT IMPLEMENT UNTIL PROGRAM IS READY

    """
    # final_submit = driver.find_element(By.ID,'finalSubmit')
    # final_submit.click()

    driver.close()

    time.sleep(5)

print('''
    Application Submission for all projects is now complete!
    Please review Chicago website to ensure all have been submitted successfully.
    Then, review your excel sheet to ensure the CDOT ID's have been actualized accurately.
''')
